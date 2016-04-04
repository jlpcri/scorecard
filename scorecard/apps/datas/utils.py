from openpyxl.styles import PatternFill, Font, colors, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

HEAD_QE = [
    'Week Ending',
    'Quality Innovation',
    'Staff',
    'Opening Positions (Recruiting, Pending start etc.)',
    'Contractors',
    'Throughput',
    'Backlog Story Points',
    'Story Points Prep',
    'Story Points Execution',
    'Unit Tests Dev',
    'Unit Testing Coverage (%age)',
    'Defects due to Dev',
    'Elicitation Analysis',
    'Quality',
    'SLAS met (%age) - Goal=100%',
    'UAT defects not prevented',
    'SDIs not prevented',
    'Resource swap',
    'Escalations',
    'Rework introduced (Hours)',
    'Efficiency',
    'Overtime Weekday',
    'Overtime Weekend',
    'Average Throughput = Story Points / Developers',
    'Rework Hours',
    'Resource swap (Hours)',
    'PTO',
    'Customer Facing Time',
    'Documentation Time',
    'Ticketless Dev Time',
    'Costs',
    'Coding Operational Costs',
    'Licensing Costs',
    'Total Operational Costs',
    'External Savings',
    'Internal Savings',
    'Usage',
    'Pheme manual tests',
    'Pheme automatic tests',
    'Visilog TXL parsed',
    'Visilog TXL schema violation found',
    'CEEQ Daily Summaries',
    'Other savings'
]

HEAD_RE = [
    'Week Ending',
    'Requirement Engineering',
    'Staff',
    'Opening Positions (Recruiting, Pending start etc.)',
    'Contractors',
    'Throughput',
    'Backlog',
    'Active Projects',
    'Team Initiatives',
    'Elicitation and Analysis Hours',
    'Quality',
    'Revisions (post review)',
    'Rework introduced (Hours)',
    'Avg Throughput = Active Projects + Team initiatives',
    'SLAs met (%age)',
    'SLAs missed (%age)',
    'Escalations',
    'Efficiency',
    'PTO Holiday',
    'Gross Available Hours',
    'Efficiency - (Elicitation & Analysis hrs)/Gross Available',
    'Overtime Weekend in Hours (from PARTE)',
    'Overtime Weekday Hours (greater than 6 hours /day)',
    'Rework From External teams (Hours)',
    'Costs',
    'Cost of Rework',
    'Resource Swap (Hours)',
    'Operational Costs',
    'Travel Costs',
    'Overall Operating Costs',
    'Other Savings'
]

HEAD_TL = [
    'Week Ending',
    'Test Lab',
    'Staff',
    'Opening Positions (Recruiting, Pending start etc.)',
    'Contractors',
    'Throughput',
    'Tickets Received',
    'Tickets Closed',
    'Virtual Machines',
    'Physical Machines',
    'Monitor Machines',
    'Quality',
    'Builds Submitted',
    'Builds Accepted',
    'Builds Rejected',
    'Platform Drift Violations',
    'Updates to Install Docs',
    'Efficiency',
    'Administration Time',
    'Project Time',
    'Ticket Time',
    'SLAs Met',
    'PTO Holiday',
    'Utilization',
    'Efficiency',
    'Costs',
    'Power Consumption - UPS A(KW)',
    'Power Consumption - UPS B(KW)',
    'Licensing Costs',
    'Other Savings'
]

HEAD_QA = [
    'Week Ending',
    'Quality Assurance',
]

HEAD_TE = [
    'Week Ending',
    'Test Engineering',
]

HEAD_TESTING = [
    'Staff',
    'Opening Positions (Recruiting, Pending start etc.)',
    'Contractors',
    'Throughput',
    'Team Initiatives',
    'Backlog Tickets',
    'Tickets Prep',
    'Tickets Execution',
    'Tickets Closure',
    'Backlog Projects',
    'Projects Prep',
    'Projects Execution',
    'Projects Closure',
    'Manual TCs Dev',
    'Manual TC dev(Hours)',
    'Automated TCs Dev',
    'Automated TC Dev(Hours)',
    'Automation Footprint Dev %age',
    'Manual TCs Execution',
    'Manual TCs Execution (Hours)',
    'Automated TCs Execution',
    'Automated TCs Execution (Hours)',
    'Automation Footprint Execution %age',
    'Average Throughput = (TCs dev + TCs Exec)/testers',
    'Quality',
    'SLAs met (%age) - Goal=100%',
    'LOE Deviation (Hours)',
    'Defects Caught',
    'UAT defects not prevented',
    'SDIs not prevented',
    'Resource Swap',
    'Escalations',
    'Standard violated',
    'Rework Introduced (Hours)',
    'Efficiency',
    'PTO Holiday (Hours)',
    'Productive Time (Hours)',
    'Active Tickets',
    'Active Projects',
    'Automation + Execution (Hours)',
    'Gross Available (Hours)',
    'Efficiency - (Automation + Execution hrs)/Gross Available',
    'Overtime Weekend (Hours)',
    'Overtime Weekday (Hours)',
    'Rework Hours',
    'Resource Swap (Hours)',
    'Costs',
    'Testing Operational Costs',
    'Licensing Costs',
    'Total Operational Costs',
    'Cost Saved by Automation',
    'Other savings'
]

HEAD_TESTING_SUMMARY = [
    'Week Ending',
    'Testing Combined',
    'Staff',
    'Opening Positions (Recruiting, Pending start etc.)',
    'Contractors',
    'Throughput',
    'Team Initiatives',
    'Active Projects',
    'Active Tickets',
    'Avg Throughput per week',
    'Quality',
    'Delays Introduced',
    'Defects identified',
    'UAT defects not prevented',
    'SDIs not prevented',
    'Average team size',
    'Avg timeframe (project)',
    'Automation Footprint',
    'Automation Hours',
    'Test Execution Hours',
    'Gross Available Hours',
    'Efficiency',
    'Overtime Weekend in Hours',
    'Overtime Weekday Hours',
    'Rework Hours',
    'Resource Swap (hours)',
    'Escalations',
    'Overall Operating Costs',
    'Total Savings'
]

HEAD_QE_TL_SUMMARY = [
    'Week Ending',
    'Quality Innovation',
    'Staff',
    'Opening Positions (Recruiting, Pending start etc.)',
    'Contractors',
    'Throughput',
    'Active Projects',
    'Elicitation and Analysis Hours',
    'Rework Hours',
    'Avg Throughput per developer',
    'Delays introduced (Hours)',
    '# of revisions',
    'Costs',
    'Tests executed on platform',
    'Savings',
    'Value Add',
    'Test Lab',
    'Staff',
    'Opening Positions (Recruiting, Pending start etc.)',
    'Contractors',
    'Throughput',
    'Tickets Received',
    'Tickets Closed',
    'Virtual Machines',
    'Physical machines',
    'Costs',
    'Power Consumption - UPS A(kw)',
    'Power Consumption - UPS B(kw)',
    'Licensing Costs'
]

blackFill = PatternFill(
    fill_type='solid',
    start_color='FF000000',
    end_color='FFFFFFFF')
cornFlowerBlueFill = PatternFill(
    start_color='4F81BD',
    end_color='4F81BD',
    fill_type='solid')
oliveDrabFill = PatternFill(
    start_color='9BBB59',
    end_color='9BBB59',
    fill_type='solid')
rebeccaPurpleFill = PatternFill(
    start_color='8064A2',
    end_color='8064A2',
    fill_type='solid')
saddleBrownFill = PatternFill(
    start_color='F79646',
    end_color='F79646',
    fill_type='solid')
darkGreenFill = PatternFill(
    start_color='00B050',
    end_color='00B050',
    fill_type='solid')
paleGreenFill = PatternFill(
    start_color='D8E4BC',
    end_color='D8E4BC',
    fill_type='solid')
silverGrayFill = PatternFill(
    start_color='C0C0C0',
    end_color='C0C0C0',
    fill_type='solid')
cornSilkFill = PatternFill(
    start_color='EEECE1',
    end_color='EEECE1',
    fill_type='solid')


headAlignment = Alignment(horizontal='center',
                          vertical='bottom',
                          textRotation=60,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)

thinBorder = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

mediumBorder = Border(left=Side(style='medium'),
                      right=Side(style='medium'),
                      top=Side(style='medium'),
                      bottom=Side(style='medium'))

ROW_START_INDEX = 4

COL_EXCLUDE_QE = [2, 6, 14, 21, 31, 37]
COL_EXCLUDE_RE = [2, 6, 11, 18, 25]
COL_EXCLUDE_TL = [2, 6, 12, 18, 26]
COL_EXCLUDE_QA_TE = [2, 6, 27, 37, 49]
COL_EXCLUDE_TEST_SUMMARY = [2, 6, 11]
COL_EXCLUDE_QE_TL_SUMMARY = [2, 6, 13, 17, 21, 26]


def write_to_excel(metric, ws):
    row_start = ROW_START_INDEX

    if metric.functional_group.abbreviation == 'QE':
        # write head title
        write_head_title(ws, HEAD_QE)

        # apply header style
        apply_header_style(ws, len(HEAD_QE) + 1, metric.functional_group.abbreviation)

        # apply border style
        apply_border_style(ws, row_start + 1, len(HEAD_QE) + 1)

        write_ytd(ws, len(HEAD_QE) + 1, row_start, row_start, COL_EXCLUDE_QE)

        write_body_qi(ws, row_start, metric)

    elif metric.functional_group.abbreviation == 'TL':
        # write head title
        write_head_title(ws, HEAD_TL)

        # apply border style
        apply_border_style(ws, row_start + 1, len(HEAD_TL) + 1)

        write_ytd(ws, len(HEAD_TL) + 1, row_start, row_start, COL_EXCLUDE_TL)

        write_body_tl(ws, row_start, metric)

    elif metric.functional_group.abbreviation == 'RE':
        write_head_title(ws, HEAD_RE)

        apply_header_style(ws, len(HEAD_RE) + 1, metric.functional_group.abbreviation)

        apply_border_style(ws, row_start + 1, len(HEAD_RE) + 1)

        write_ytd(ws, len(HEAD_RE) + 1, row_start, row_start, COL_EXCLUDE_RE)

        write_body_re(ws, row_start, metric)

    else:
        head = []
        if metric.functional_group.abbreviation == 'QA':
            head = HEAD_QA + HEAD_TESTING
        elif metric.functional_group.abbreviation == 'TE':
            head = HEAD_TE + HEAD_TESTING

        write_head_title(ws, head)

        apply_header_style(ws, len(head) + 1, metric.functional_group.abbreviation)

        apply_border_style(ws, row_start + 1, len(head) + 1)

        write_ytd(ws, len(head) + 1, row_start, row_start, COL_EXCLUDE_QA_TE)

        write_body_qa_te(ws, row_start, metric)


def write_head_title(ws, title):
    for col in range(1, len(title) + 1):
        cell = ws.cell(row=1, column=col)
        cell.value = title[col - 1]
        cell.alignment = headAlignment
        cell.font = Font(bold=True, size=12)


def apply_border_style(ws, rows, columns):
    for r_index in range(1, rows):
        for c_index in range(1, columns):
            ws.cell(row=r_index, column=c_index).border = thinBorder
            ws.cell(row=r_index, column=c_index).number_format = '0.00'

    for c_index in range(2, columns):
        ws.column_dimensions[get_column_letter(c_index)].width = 10


def write_body_qi(ws, row, metric):
    # column 1-6
        write_human_resource(ws, row, metric)

        ws.cell(row=row, column=7).value = metric.story_points_backlog
        ws.cell(row=row, column=8).value = metric.story_points_prep
        ws.cell(row=row, column=9).value = metric.story_points_execution
        ws.cell(row=row, column=10).value = metric.unit_tests_dev
        ws.cell(row=row, column=11).value = metric.unit_tests_coverage
        ws.cell(row=row, column=11).number_format = '0.00%'
        ws.cell(row=row, column=12).value = metric.defects_in_dev
        ws.cell(row=row, column=13).value = metric.elicitation_analysis_time

        # column 14
        background_color_fill(ws, row, col=14, background_color=rebeccaPurpleFill)
        ws.column_dimensions[get_column_letter(14)].width = 4

        ws.cell(row=row, column=15).value = metric.slas_met
        ws.cell(row=row, column=16).value = metric.uat_defects_not_prevented
        ws.cell(row=row, column=17).value = metric.sdis_not_prevented
        ws.cell(row=row, column=18).value = metric.resource_swap
        ws.cell(row=row, column=19).value = metric.escalations
        ws.cell(row=row, column=20).value = metric.rework_introduced_time

        # column 21
        background_color_fill(ws, row, col=21, background_color=saddleBrownFill)
        ws.column_dimensions[get_column_letter(21)].width = 4

        ws.cell(row=row, column=22).value = metric.overtime_weekday
        ws.cell(row=row, column=23).value = metric.overtime_weekend
        ws.cell(row=row, column=24).value = metric.avg_throughput
        ws.cell(row=row, column=25).value = metric.rework_time
        ws.cell(row=row, column=26).value = metric.resource_swap_time
        ws.cell(row=row, column=27).value = metric.pto_holiday_time
        ws.cell(row=row, column=28).value = metric.customer_facing_time
        ws.cell(row=row, column=29).value = metric.documentation_time
        ws.cell(row=row, column=30).value = metric.ticketless_dev_time

        # column 31
        background_color_fill(ws, row, col=31, background_color=oliveDrabFill)
        ws.column_dimensions[get_column_letter(31)].width = 4

        # add $ symbol
        add_dollar_symbol(ws, row, col_start=32, col_end=36)

        ws.cell(row=row, column=32).value = metric.operational_cost
        ws.cell(row=row, column=33).value = metric.license_cost
        ws.cell(row=row, column=34).value = metric.total_operational_cost
        ws.cell(row=row, column=35).value = metric.external_savings
        ws.cell(row=row, column=36).value = metric.internal_savings

        # column 37
        background_color_fill(ws, row, col=37, background_color=darkGreenFill)
        ws.column_dimensions[get_column_letter(37)].width = 4

        ws.cell(row=row, column=38).value = metric.pheme_manual_tests
        ws.cell(row=row, column=39).value = metric.pheme_auto_tests
        ws.cell(row=row, column=40).value = metric.visilog_txl_parsed
        ws.cell(row=row, column=41).value = metric.visilog_txl_schema_violation
        ws.cell(row=row, column=42).value = metric.ceeq_daily_summaries

        add_dollar_symbol(ws, row, col_start=43, col_end=43)
        ws.cell(row=row, column=43).value = metric.other_savings


def write_body_tl(ws, row, metric):
    # column 1-6
    write_human_resource(ws, row, metric)

    ws.cell(row=row, column=7).value = metric.tickets_received
    ws.cell(row=row, column=8).value = metric.tickets_closed
    ws.cell(row=row, column=9).value = metric.virtual_machines
    ws.cell(row=row, column=10).value = metric.physical_machines
    ws.cell(row=row, column=11).value = metric.monitor_machines

    # column 12
    background_color_fill(ws, row, col=12, background_color=rebeccaPurpleFill)
    ws.column_dimensions[get_column_letter(12)].width = 4

    # column 13
    ws.cell(row=row, column=13).value = metric.builds_submitted
    ws.cell(row=row, column=14).value = metric.builds_accepted
    ws.cell(row=row, column=15).value = metric.builds_rejected
    ws.cell(row=row, column=16).value = metric.platform_drift_violations
    ws.cell(row=row, column=17).value = metric.updates_install_docs

    # column 18
    background_color_fill(ws, row, col=18, background_color=saddleBrownFill)
    ws.column_dimensions[get_column_letter(18)].width = 4

    # column 19
    ws.cell(row=row, column=19).value = metric.administration_time
    ws.cell(row=row, column=20).value = metric.project_time
    ws.cell(row=row, column=21).value = metric.ticket_time
    ws.cell(row=row, column=22).value = metric.slas_met
    ws.cell(row=row, column=23).value = metric.pto_holiday_time
    ws.cell(row=row, column=24).value = metric.utilization
    ws.cell(row=row, column=25).value = metric.efficiency

    # column 26
    background_color_fill(ws, row, col=26, background_color=darkGreenFill)
    ws.column_dimensions[get_column_letter(26)].width = 4

    ws.cell(row=row, column=27).value = metric.power_consumption_ups_a
    ws.cell(row=row, column=28).value = metric.power_consumption_ups_b
    ws.cell(row=row, column=29).value = metric.license_cost
    ws.cell(row=row, column=30).value = metric.other_savings

    add_dollar_symbol(ws, row, col_start=29, col_end=30)


def write_body_re(ws, row, metric):
    # column 1-6
    write_human_resource(ws, row, metric)

    # column 7
    ws.cell(row=row, column=7).value = metric.backlog
    ws.cell(row=row, column=8).value = metric.active_projects
    ws.cell(row=row, column=9).value = metric.team_initiative
    ws.cell(row=row, column=10).value = metric.elicitation_analysis_time

    # column 11
    background_color_fill(ws, row, col=11, background_color=rebeccaPurpleFill)
    ws.column_dimensions[get_column_letter(11)].width = 4

    # column 12
    ws.cell(row=row, column=12).value = metric.revisions
    ws.cell(row=row, column=13).value = metric.rework_introduced_time
    ws.cell(row=row, column=14).value = metric.avg_throughput
    ws.cell(row=row, column=15).value = metric.slas_met
    ws.cell(row=row, column=16).value = metric.slas_missed
    # ws.cell(row=row, column=17).value = metric.delays_introduced_time
    ws.cell(row=row, column=17).value = metric.escalations

    # column 18
    background_color_fill(ws, row, col=18, background_color=saddleBrownFill)
    ws.column_dimensions[get_column_letter(18)].width = 4

    # column 19
    ws.cell(row=row, column=19).value = metric.pto_holiday_time
    ws.cell(row=row, column=20).value = metric.gross_available_time
    ws.cell(row=row, column=21).value = metric.efficiency
    ws.cell(row=row, column=21).number_format = '0.00%'
    ws.cell(row=row, column=22).value = metric.overtime_weekend
    ws.cell(row=row, column=23).value = metric.overtime_weekday
    ws.cell(row=row, column=24).value = metric.rework_external_time

    # column 25
    background_color_fill(ws, row, col=25, background_color=oliveDrabFill)
    ws.column_dimensions[get_column_letter(25)].width = 4

    # column 26
    add_dollar_symbol(ws, row, col_start=26, col_end=26)
    add_dollar_symbol(ws, row, col_start=28, col_end=31)

    ws.cell(row=row, column=26).value = metric.rework_external_cost
    ws.cell(row=row, column=27).value = metric.resource_swap_time
    ws.cell(row=row, column=28).value = metric.operational_cost
    ws.cell(row=row, column=29).value = metric.travel_cost
    ws.cell(row=row, column=30).value = metric.overall_cost
    ws.cell(row=row, column=31).value = metric.other_savings


def write_body_qa_te(ws, row, metric):
    # column 1-6
    write_human_resource(ws, row, metric)

    # column 7
    ws.cell(row=row, column=7).value = metric.team_initiative
    ws.cell(row=row, column=8).value = metric.ticket_backlog
    ws.cell(row=row, column=9).value = metric.ticket_prep
    ws.cell(row=row, column=10).value = metric.ticket_execution
    ws.cell(row=row, column=11).value = metric.ticket_closed
    ws.cell(row=row, column=12).value = metric.project_backlog
    ws.cell(row=row, column=13).value = metric.project_prep
    ws.cell(row=row, column=14).value = metric.project_execution
    ws.cell(row=row, column=15).value = metric.project_closed
    ws.cell(row=row, column=16).value = metric.tc_manual_dev
    ws.cell(row=row, column=17).value = metric.tc_manual_dev_time
    ws.cell(row=row, column=18).value = metric.tc_auto_dev
    ws.cell(row=row, column=19).value = metric.tc_auto_dev_time

    ws.cell(row=row, column=20).value = metric.auto_footprint_dev_age
    ws.cell(row=row, column=20).number_format = '0.00%'

    ws.cell(row=row, column=21).value = metric.tc_manual_execution
    ws.cell(row=row, column=22).value = metric.tc_manual_execution_time
    ws.cell(row=row, column=23).value = metric.tc_auto_execution
    ws.cell(row=row, column=24).value = metric.tc_auto_execution_time

    ws.cell(row=row, column=25).value = metric.auto_footprint_execution_age
    ws.cell(row=row, column=25).number_format = '0.00%'

    ws.cell(row=row, column=26).value = metric.avg_throughput

    # column 27
    background_color_fill(ws, row, col=27, background_color=rebeccaPurpleFill)
    ws.column_dimensions[get_column_letter(27)].width = 4

    # column 28
    ws.cell(row=row, column=28).value = metric.slas_met
    ws.cell(row=row, column=29).value = metric.loe_deviation
    ws.cell(row=row, column=30).value = metric.defect_caught
    ws.cell(row=row, column=31).value = metric.uat_defects_not_prevented
    ws.cell(row=row, column=32).value = metric.sdis_not_prevented
    ws.cell(row=row, column=33).value = metric.resource_swap
    ws.cell(row=row, column=34).value = metric.escalations
    ws.cell(row=row, column=35).value = metric.standards_violated
    ws.cell(row=row, column=36).value = metric.rework_introduced_time

    # column 37
    background_color_fill(ws, row, col=37, background_color=saddleBrownFill)
    ws.column_dimensions[get_column_letter(37)].width = 4

    # column 38
    ws.cell(row=row, column=38).value = metric.pto_holiday_time
    ws.cell(row=row, column=39).value = metric.productive_hours
    ws.cell(row=row, column=40).value = metric.active_tickets
    ws.cell(row=row, column=41).value = metric.active_projects
    ws.cell(row=row, column=42).value = metric.auto_and_execution_time
    ws.cell(row=row, column=43).value = metric.gross_available_time

    ws.cell(row=row, column=44).value = metric.efficiency
    ws.cell(row=row, column=44).number_format = '0.00%'

    ws.cell(row=row, column=45).value = metric.overtime_weekend
    ws.cell(row=row, column=46).value = metric.overtime_weekday
    ws.cell(row=row, column=47).value = metric.rework_time
    ws.cell(row=row, column=48).value = metric.resource_swap_time

    # column 49
    background_color_fill(ws, row, col=49, background_color=oliveDrabFill)
    ws.column_dimensions[get_column_letter(49)].width = 4

    # column 50
    add_dollar_symbol(ws, row, col_start=50, col_end=54)
    ws.cell(row=row, column=50).value = metric.operational_cost
    ws.cell(row=row, column=51).value = metric.license_cost
    ws.cell(row=row, column=52).value = metric.total_operational_cost
    ws.cell(row=row, column=53).value = metric.auto_savings
    ws.cell(row=row, column=54).value = metric.other_savings


def write_human_resource(ws, row, metric):
    # column 1 Weed Ending
    ws.cell(row=row, column=1).value = get_week_ending_date(metric.created)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')

    # column 2
    background_color_fill(ws, row, col=2, background_color=blackFill)
    ws.column_dimensions[get_column_letter(2)].width = 4

    ws.cell(row=row, column=3).value = metric.staffs
    ws.cell(row=row, column=4).value = metric.openings
    ws.cell(row=row, column=5).value = metric.contractors

    # column 6
    background_color_fill(ws, row, col=6, background_color=cornFlowerBlueFill)
    ws.column_dimensions[get_column_letter(6)].width = 4


def get_week_ending_date(date):
    if date.month < 10:
        month = '0' + str(date.month)
    else:
        month = str(date.month)

    if date.day < 10:
        day = '0' + str(date.day)
    else:
        day = str(date.day)

    return str(date.year) + '-' + month + '-' + day


def background_color_fill(ws, row, col, background_color):
    ws.cell(row=1, column=col).font = Font(color=colors.WHITE, size=12, bold=True)
    for r_index in range(1, row + 1):
        ws.cell(row=r_index, column=col).fill = background_color


def add_dollar_symbol(ws, row, col_start, col_end):
    for r_index in range(2, row + 1):
        for c_index in range(col_start, col_end + 1):
            ws.cell(row=r_index, column=c_index).number_format = '$0.00'


def write_to_excel_all(metrics, ws, key):
    row_start = ROW_START_INDEX

    if key == 'QE':
        write_head_title(ws, HEAD_QE)

        apply_header_style(ws, len(HEAD_QE) + 1, key)

        apply_border_style(ws, row_start + len(metrics), len(HEAD_QE) + 1)

        write_ytd(ws, len(HEAD_QE) + 1, row_start, row_start + len(metrics) - 1, COL_EXCLUDE_QE)

        for metric in metrics:
            write_body_qi(ws, row_start, metric)
            row_start += 1
    elif key == 'TL':
        write_head_title(ws, HEAD_TL)

        apply_border_style(ws, row_start + len(metrics), len(HEAD_TL) + 1)

        write_ytd(ws, len(HEAD_TL) + 1, row_start, row_start + len(metrics) - 1, COL_EXCLUDE_TL)

        for metric in metrics:
            write_body_tl(ws, row_start, metric)
            row_start += 1

    elif key == 'RE':
        write_head_title(ws, HEAD_RE)

        apply_header_style(ws, len(HEAD_RE) + 1, key)

        apply_border_style(ws, row_start + len(metrics), len(HEAD_RE) + 1)

        write_ytd(ws, len(HEAD_RE) + 1, row_start, row_start + len(metrics) - 1, COL_EXCLUDE_RE)

        for metric in metrics:
            write_body_re(ws, row_start, metric)
            row_start += 1

    else:
        head = []
        if key == 'QA':
            head = HEAD_QA + HEAD_TESTING
        elif key == 'TE':
            head = HEAD_TE + HEAD_TESTING

        write_head_title(ws, head)

        apply_header_style(ws, len(head) + 1, key)

        apply_border_style(ws, row_start + len(metrics), len(head) + 1)

        write_ytd(ws, len(head) + 1, row_start, row_start + len(metrics) - 1, COL_EXCLUDE_QA_TE)

        for metric in metrics:
            write_body_qa_te(ws, row_start, metric)
            row_start += 1


def write_ytd(ws, columns, row_start, row_end, col_exclude):
    ws.column_dimensions['A'].width = 12
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='right')

    ws.cell(row=2, column=1).value = 'YTD Average'
    ws.cell(row=3, column=1).value = 'YTD Totals'

    for r_index in range(2, 4):
        for c_index in range(1, columns):
            ws.cell(row=r_index, column=c_index).border = mediumBorder
            ws.cell(row=r_index, column=c_index).font = Font(bold=True)
            if c_index > 1:
                ws.cell(row=r_index, column=c_index).fill = silverGrayFill

    for c_index in range(2, columns):
        if c_index in col_exclude:
            continue
        c_letter = get_column_letter(c_index)
        start = c_letter + str(row_start)
        end = c_letter + str(row_end)
        # print start, end
        ws.cell(row=2, column=c_index).value = '=AVERAGE({0}:{1})'.format(start, end)
        ws.cell(row=3, column=c_index).value = '=SUM({0}:{1})'.format(start, end)


def write_to_excel_test_summary(ws, dates):
    row_start = ROW_START_INDEX

    write_head_title(ws, HEAD_TESTING_SUMMARY)

    apply_header_style(ws, len(HEAD_TESTING_SUMMARY) + 1, 'TEST_SUMMARY')

    apply_border_style(ws, row_start + len(dates), len(HEAD_TESTING_SUMMARY) + 1)

    write_ytd(ws, len(HEAD_TESTING_SUMMARY) + 1, row_start, row_start + len(dates) - 1, COL_EXCLUDE_TEST_SUMMARY)

    for date in dates:
        # column 1
        ws.cell(row=row_start, column=1).value = get_week_ending_date(date)

        # column 2
        background_color_fill(ws, row_start, col=2, background_color=blackFill)
        ws.column_dimensions[get_column_letter(2)].width = 4

        # column 3
        ws.cell(row=row_start, column=3).value = get_formula_sum_from_qa_te(row_start, 3)  # staff
        ws.cell(row=row_start, column=4).value = get_formula_sum_from_qa_te(row_start, 4)  # openings
        ws.cell(row=row_start, column=5).value = get_formula_sum_from_qa_te(row_start, 5)  # contractors

        # column 6
        background_color_fill(ws, row_start, col=6, background_color=cornFlowerBlueFill)
        ws.column_dimensions[get_column_letter(6)].width = 4

        # column 7
        ws.cell(row=row_start, column=7).value = get_formula_sum_from_qa_te(row_start, 7)    # team initiatives
        ws.cell(row=row_start, column=8).value = get_formula_sum_from_qa_te(row_start, 41)   # active projects
        ws.cell(row=row_start, column=9).value = get_formula_sum_from_qa_te(row_start, 40)   # active tickets
        ws.cell(row=row_start, column=10).value = get_formula_avg_from_qa_te(row_start, 26)  # avg throughput

        # column 11
        background_color_fill(ws, row_start, col=11, background_color=rebeccaPurpleFill)
        ws.column_dimensions[get_column_letter(11)].width = 4

        # column 12
        ws.cell(row=row_start, column=12).value = get_formula_sum_from_qa_te(row_start, 29)  # delays introduced
        ws.cell(row=row_start, column=13).value = get_formula_sum_from_qa_te(row_start, 30)  # defects identified
        ws.cell(row=row_start, column=14).value = get_formula_sum_from_qa_te(row_start, 31)  # uat defects not prevented
        ws.cell(row=row_start, column=15).value = get_formula_sum_from_qa_te(row_start, 32)  # SDIs not prevented
        ws.cell(row=row_start, column=16).value = get_formula_avg_from_qa_te(row_start, 38)  # Avg team size
        ws.cell(row=row_start, column=17).value = get_formula_avg_from_qa_te(row_start, 39)  # avg timeframe

        ws.cell(row=row_start, column=18).value = get_formula_sum_from_qa_te(row_start, 20)  # automation footprint
        ws.cell(row=row_start, column=18).number_format = '0.00%'

        ws.cell(row=row_start, column=19).value = get_formula_sum_from_qa_te(row_start, 24)  # automation hours
        ws.cell(row=row_start, column=20).value = get_formula_sum_from_qa_te(row_start, 22)  # test exec hours
        ws.cell(row=row_start, column=21).value = get_formula_sum_from_qa_te(row_start, 43)  # gross available
        ws.cell(row=row_start, column=22).value = get_formula_avg_from_qa_te(row_start, 44)  # efficiency
        ws.cell(row=row_start, column=22).number_format = '0.00%'
        ws.cell(row=row_start, column=23).value = get_formula_sum_from_qa_te(row_start, 45)  # overtime weekend
        ws.cell(row=row_start, column=24).value = get_formula_sum_from_qa_te(row_start, 46)  # overtime weekday
        ws.cell(row=row_start, column=25).value = get_formula_sum_from_qa_te(row_start, 47)  # rework hours
        ws.cell(row=row_start, column=26).value = get_formula_sum_from_qa_te(row_start, 48)  # resource swap hours
        ws.cell(row=row_start, column=27).value = get_formula_sum_from_qa_te(row_start, 34)  # escalations
        ws.cell(row=row_start, column=28).value = get_formula_sum_from_qa_te(row_start, 52)  # overall opening costs
        ws.cell(row=row_start, column=29).value = get_formula_sum_from_qa_te_total_savings(row_start, [53, 54])  # total savings

        row_start += 1

    add_dollar_symbol(ws, row_start, col_start=28, col_end=29)
    ws.column_dimensions[get_column_letter(28)].width = 14
    ws.column_dimensions[get_column_letter(29)].width = 14


def write_to_excel_qi_tl_summary(ws, dates):
    row_start = ROW_START_INDEX

    write_head_title(ws, HEAD_QE_TL_SUMMARY)

    apply_header_style(ws, len(HEAD_QE_TL_SUMMARY) + 1, 'QE_TL_SUMMARY')

    apply_border_style(ws, row_start + len(dates), len(HEAD_QE_TL_SUMMARY) + 1)

    write_ytd(ws, len(HEAD_QE_TL_SUMMARY) + 1, row_start, row_start + len(dates) - 1, COL_EXCLUDE_QE_TL_SUMMARY)

    for date in dates:
        # column 1
        ws.cell(row=row_start, column=1).value = get_week_ending_date(date)

        # column 2
        background_color_fill(ws, row_start, col=2, background_color=blackFill)
        ws.column_dimensions[get_column_letter(2)].width = 4

        # column 3
        ws.cell(row=row_start, column=3).value = get_cell_value_from_qi(row_start, 3)
        ws.cell(row=row_start, column=4).value = get_cell_value_from_qi(row_start, 4)
        ws.cell(row=row_start, column=5).value = get_cell_value_from_qi(row_start, 5)

        # column 6
        background_color_fill(ws, row_start, col=6, background_color=cornFlowerBlueFill)
        ws.column_dimensions[get_column_letter(6)].width = 4

        # column 7
        ws.cell(row=row_start, column=7).value = get_cell_value_from_qi(row_start, 14)   # active projects
        ws.cell(row=row_start, column=8).value = get_cell_value_from_qi(row_start, 15)   # elicitation analysis
        ws.cell(row=row_start, column=9).value = get_cell_value_from_qi(row_start, 30)   # rework hours
        ws.cell(row=row_start, column=10).value = get_cell_value_from_qi(row_start, 29)  # avg throughput
        ws.cell(row=row_start, column=11).value = get_cell_value_from_qi(row_start, 19)  # delays introduced hours
        ws.cell(row=row_start, column=12).value = get_cell_value_from_qi(row_start, 16)  # revisions

        # column 13
        background_color_fill(ws, row_start, col=13, background_color=oliveDrabFill)
        ws.column_dimensions[get_column_letter(13)].width = 4

        # column 14
        ws.cell(row=row_start, column=14).value = get_formula_sum_from_qi(row_start, 37)
        ws.cell(row=row_start, column=15).value = get_formula_savings_from_qi(row_start, 37)
        ws.cell(row=row_start, column=16).value = get_formula_value_add_from_qi(row_start, 39)

        # column 17
        background_color_fill(ws, row_start, col=17, background_color=blackFill)
        ws.column_dimensions[get_column_letter(17)].width = 4

        # column 18
        ws.cell(row=row_start, column=18).value = get_cell_value_from_tl(row_start, 3)
        ws.cell(row=row_start, column=19).value = get_cell_value_from_tl(row_start, 4)
        ws.cell(row=row_start, column=20).value = get_cell_value_from_tl(row_start, 5)

        # column 21
        background_color_fill(ws, row_start, col=21, background_color=cornFlowerBlueFill)
        ws.column_dimensions[get_column_letter(21)].width = 4

        # column 22
        ws.cell(row=row_start, column=22).value = get_cell_value_from_tl(row_start, 7)
        ws.cell(row=row_start, column=23).value = get_cell_value_from_tl(row_start, 8)
        ws.cell(row=row_start, column=24).value = get_cell_value_from_tl(row_start, 9)
        ws.cell(row=row_start, column=25).value = get_cell_value_from_tl(row_start, 10)

        # column 26
        background_color_fill(ws, row_start, col=26, background_color=oliveDrabFill)
        ws.column_dimensions[get_column_letter(26)].width = 4

        # column 27
        ws.cell(row=row_start, column=27).value = get_cell_value_from_tl(row_start, 12)
        ws.cell(row=row_start, column=28).value = get_cell_value_from_tl(row_start, 13)
        ws.cell(row=row_start, column=29).value = get_cell_value_from_tl(row_start, 14)

        row_start += 1

    add_dollar_symbol(ws, row_start, col_start=29, col_end=29)


def get_formula_sum_from_qa_te(row, col):
    row = str(row)
    return '=SUM(\'Quality Assurance\'!{0}, \'Test Engineering\'!{0})'.format(get_column_letter(col) + row)


def get_formula_sum_from_qa_te_total_savings(row, cols):
    row = str(row)
    string = ''
    for col in cols:
        tmp = '\'Quality Assurance\'!{0}, \'Test Engineering\'!{0},'.format(get_column_letter(col) + row)
        string += tmp

    string = '=SUM(' + string + ')'
    return string


def get_formula_avg_from_qa_te(row, col):
    row = str(row)
    return '=AVERAGE(\'Quality Assurance\'!{0}, \'Test Engineering\'!{0})'.format(get_column_letter(col) + row)


def get_cell_value_from_qi(row, col):
    row = str(row)

    return '=\'Quality Engineering\'!{0}'.format(get_column_letter(col) + row)


def get_formula_sum_from_qi(row, col):
    row = str(row)

    return '=SUM(\'Quality Engineering\'!{0}:{1})'.format(get_column_letter(col) + row, get_column_letter(col + 2) + row)


def get_formula_savings_from_qi(row, col):
    row = str(row)

    return '=(\'Quality Engineering\'!{0}*1.73)+(\'Quality Engineering\'!{1}*1.97)'.format(get_column_letter(col) + row, get_column_letter(col + 1) + row)


def get_formula_value_add_from_qi(row, col):
    row = str(row)

    return '=(\'Quality Engineering\'!{0}*0.33)'.format(get_column_letter(col) + row)


def get_cell_value_from_tl(row, col):
    row = str(row)

    return '=\'Test Lab\'!{0}'.format(get_column_letter(col) + row)


def apply_header_style(ws, cols, key):
    pale_green = []
    corn_silk = []
    if key == 'QE':
        pale_green = [4, 11, 12, 19, 26, 34, 37, 38, 39, 40, 41]
        corn_silk = [33, 35]
    elif key == 'RE':
        pale_green = [4, 28]
        corn_silk = [20, 21, 25, 27, 29]
    elif key in ['QA', 'TE']:
        pale_green = [4, 20, 24, 38, 51, 54]
        corn_silk = [25, 26, 42, 43, 44, 50, 52, 53]
    elif key == 'TEST_SUMMARY':
        pale_green = [4, 10, 18, 22, 28, 29]
        corn_silk = [23, 24]
    elif key == 'QE_TL_SUMMARY':
        pale_green = [4, 19, 24, 25, 27, 28, 29]
        corn_silk = [10]

    for col in range(1, cols):
        if col in pale_green:
            ws.cell(row=1, column=col).fill = paleGreenFill
        elif col in corn_silk:
            ws.cell(row=1, column=col).fill = cornSilkFill
