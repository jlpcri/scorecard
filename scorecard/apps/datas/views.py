from datetime import datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.template import RequestContext
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from scorecard.apps.datas.utils import write_to_excel, write_to_excel_all, write_to_excel_test_summary, \
    write_to_excel_qi_tl_summary, get_week_ending_date

from scorecard.apps.users.models import FunctionalGroup
from scorecard.apps.teams.models import InnovationMetrics, LabMetrics, RequirementMetrics, TestMetrics
from scorecard.apps.teams.utils import get_start_end_from_request


@login_required
def datas(request):
    start_end = get_start_end_from_request(request)
    start = start_end['start']
    end = start_end['end']

    functional_groups = FunctionalGroup.objects.all().order_by('name')
    data = {}
    data_name = []
    data_key = []
    for functional_group in functional_groups:
        data_name.append(functional_group.name)
        data_key.append(functional_group.abbreviation)

    dates = InnovationMetrics.objects.filter(created__range=(start, end), subteam=None).values_list('created', flat=True).order_by('-created')
    data['name'] = data_name
    data['key'] = data_key
    data['dates'] = dates
    data['start'] = start
    data['end'] = end

    context = RequestContext(request, {
        'data': data
    })

    return render(request, 'datas/datas.html', context)


@login_required
def export_excel(request):
    wb = Workbook()
    functional_groups = FunctionalGroup.objects.all().order_by('name')
    key = request.GET.get('key', '')
    date = request.GET.get('date', '')

    today = get_week_ending_date(datetime.today())
    export_file_name = 'Scorecard-'
    export_date = ''

    update_error = False
    update_error_list = []

    start_end = get_start_end_from_request(request)
    start = start_end['start']
    end = start_end['end']

    if date:
        date = datetime.strptime(date, '%b. %d, %Y')
        # print date

        if key == 'SU':
            wb.active.title = 'Testing Summary'
            for functional_group in functional_groups:
                ws = wb.create_sheet(functional_group.name)

                metric = functional_group.metrics_set.get(subteam=None,
                                                          created__year=date.year,
                                                          created__month=date.month,
                                                          created__day=date.day)
                # if not metric.updated:
                #     update_error = True
                #     update_error_list.append(str(functional_group.abbreviation))
                # else:
                write_to_excel(metric, ws)
                export_date = metric.created

            if export_date:
                export_file_name += get_week_ending_date(export_date)

            if update_error:
                messages.error(request, 'Team {0} not updated'.format(update_error_list))
                return redirect('datas:datas')

        else:
            ws = wb.active
            ws.title = key
            # ws.sheet_properties.tabColor = '1072BA'
            group = FunctionalGroup.objects.get(abbreviation=key)
            metric = group.metrics_set.get(subteam=None,
                                           created__year=date.year,
                                           created__month=date.month,
                                           created__day=date.day)
            # print metric.id

            if not metric.updated:
                messages.error(request, 'Team \'{0}\' not updated'.format(metric.functional_group.name))
                return redirect('datas:datas')
            else:
                export_file_name += key + '-' + get_week_ending_date(metric.created)
                write_to_excel(metric, ws)

    else:
        export_file_name += 'All-' + today
        dates = InnovationMetrics.objects.filter(created__range=(start, end), subteam=None).values_list('created', flat=True).order_by('-created')

        # export formula to Testing Summar
        ws = wb.active
        ws.title = 'Testing Summary'
        write_to_excel_test_summary(ws, dates)

        # export formula to Innovation+Lab Summary
        ws = wb.create_sheet('QE and Lab Summary')
        write_to_excel_qi_tl_summary(ws, dates)

        for functional_group in functional_groups:
            ws = wb.create_sheet(functional_group.name)

            metrics = functional_group.metrics_set.filter(subteam=None,
                                                          created__range=(start, end)).order_by('-created')
            # result = check_metrics_updated(metrics)

            write_to_excel_all(metrics, ws, functional_group.abbreviation)

            # if result['valid']:
            #     write_to_excel_all(metrics, ws, functional_group.abbreviation)
            # else:
            #     update_error = True
            #     update_error_list.append(result['team'])

        # if update_error:
        #     messages.error(request, 'Team {0} not updated'.format(update_error_list))
        #     return redirect('datas:datas')

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-disposition'] = 'attachment; filename="{0}.xlsx"'.format(export_file_name)
    return response


def check_metrics_updated(metrics):
    for metric in metrics:
        if not metric.updated:
            result = {
                'valid': False,
                'team': str(metric.functional_group.abbreviation)
            }
            break
    else:
        result = {
            'valid': True
        }

    return result
