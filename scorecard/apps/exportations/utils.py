from openpyxl.styles import PatternFill, Font, colors, Alignment
from openpyxl.styles.borders import Border, Side

HEAD_QI = [
    'Week Ending',
    'Quality Innovation',
    'Staff',
    'Opening Positions',
    'Contractors',
    'Throughput',
    'Backlog Story Points',
    'Story Points Prep',
    'Story Points Execution',
    'Unit Tests Dev',
    'Unit Testing Coverage',
    'Documentation',
    'Defects due to Dev',
    'Quality',
    'SLAS met',
    'Delays Introduced',
    'UAT defects not prevented',
    'SDIs not prevented',
    'Resource swap',
    'Escalations',
    'Rework introduced',
    'Efficiency',
    'Average team size',
    'Overtime Weekday',
    'Overtime Weekend',
    'Average Throughput',
    'Rework Hours',
    'Resource swap hours',
    'Costs',
    'Coding Operational Costs',
    'Licensing Costs',
    'Total Operational Costs',
    'Usage',
    'Pheme manual tests',
    'Pheme automatic tests',
    'Visilog TXL parsed',
    'Visilog TXL schema violation found',
    'Other savings'
]

HEAD_RE = [
    'Week Ending',
    'Requirement Engineering',
    'Staff',
    'Opening Positions',
    'Contractors',
    'Throughput',
    'Backlog',
    'Active Projects',
    'Team Initiatives',
    'Elicitation and Analysis',
    'Quality',
    'Revisions',
    'Rework introduced',
    'Avg Throughput',
    'SLAs met',
    'SLAs missed',
    'Delays introduced',
    'Escalations',
    'Efficiency',
    'Gross Available Hours',
    'Efficiency',
    'Overtime Weekend',
    'Overtime Weekday',
    'Rework From External',
    'Costs',
    'Cost of Rework',
    'Resource Swap Hours',
    'Operational Costs',
    'Travel Costs',
    'Overall Operating Costs'
]

HEAD_TL = [
    'Week Ending',
    'Test Lab',
    'Staff',
    'Opening Positions',
    'Contractors',
    'Throughput',
    'Tickets Received',
    'Tickets Closed',
    'Virtual Machines',
    'Physical Machines',
    'Costs',
    'Power Consumption - UPS A(KW)',
    'Power Consumption - UPS B(KW)',
    'Licensing Costs'
]

HEAD_PQ = [
    'Week Ending',
    'Product Quality',
]

HEAD_QA = [
    'Week Ending',
    'Quality Assurance',
]

HEAD_TE = [
    'Week Ending',
    'Test Engineering',
]

HEAD_TESTING = [
    'Staff',
    'Opening Positions',
    'Contractors',
    'Throughput',
    'Team Initiatives',
    'Backlog Tickets',
    'Tickets Prep',
    'Tickets Execution',
    'Tickets Closure',
    'Backlog Projects',
    'Projects Prep',
    'Projects Execution',
    'Projects Closure',
    'Manual TCs Dev',
    'Manual TC dev(Hours)',
    'Automated TCs Dev',
    'Automated TC Dev(Hours)',
    'Automation Footprint Dev age',
    'Manual TCs Execution',
    'Manual TCs Execution (Hours)',
    'Automated TCs Execution',
    'Automated TCs Execution (Hours)',
    'Automation Footprint Execution age',
    'Average Throughput',
    'Quality',
    'SLAs met',
    'Delays Introduced (Hours)',
    'Defects Caught',
    'UAT defects not prevented',
    'SDIs not prevented',
    'Resource Swap',
    'Escalations',
    'Standard violated',
    'Rework Introduced (Hours)',
    'Efficiency',
    'Average Team size',
    'Avg timeframe (tickets)',
    'Active Tickets',
    'Active Projects',
    'Automation + Execution (Hours)',
    'Gross Available (Hours)',
    'Efficiency',
    'Overtime Weekend (Hours)',
    'Overtime Weekday (Hours)',
    'Rework Hours',
    'Resource Swap (Hours)',
    'Costs',
    'Testing Operational Costs',
    'Licensing Costs',
    'Total Operational Costs',
    'Cost Saved by Automation',
    'Other savings'
]

blackFill = PatternFill(
    fill_type='solid',
    start_color='FF000000',
    end_color='FFFFFFFF')
cornFlowerBlueFill = PatternFill(
    start_color='6495ED',
    end_color='6495ED',
    fill_type='solid')
oliveDrabFill = PatternFill(
    start_color='6B8E23',
    end_color='6B8E23',
    fill_type='solid')
rebeccaPurpleFill = PatternFill(
    start_color='663399',
    end_color='663399',
    fill_type='solid')
saddleBrownFill = PatternFill(
    start_color='8B4513',
    end_color='8B4513',
    fill_type='solid')
darkGreenFill = PatternFill(
    start_color='006400',
    end_color='006400',
    fill_type='solid')

headAlignment = Alignment(horizontal='center',
                          vertical='bottom',
                          textRotation=60,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)

thinBorder = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))


def write_to_excel(metric, ws):
    if metric.functional_group.key == 'QI':
        # write head title
        write_head_title(ws, HEAD_QI)

        row = 4

        # apply border style
        apply_border_style(ws, row + 1, len(HEAD_QI) + 1)

        # column 1-6
        write_human_resource(ws, row, metric)

        ws.cell(row=row, column=7).value = metric.story_points_backlog
        ws.cell(row=row, column=8).value = metric.story_points_prep
        ws.cell(row=row, column=9).value = metric.story_points_execution
        ws.cell(row=row, column=10).value = metric.unit_tests_dev
        ws.cell(row=row, column=11).value = metric.unit_tests_coverage
        ws.cell(row=row, column=11).number_format = '0.00%'
        ws.cell(row=row, column=12).value = metric.documentation_coverage
        ws.cell(row=row, column=12).number_format = '0.00%'
        ws.cell(row=row, column=13).value = metric.defects_in_dev

        # column 14
        background_color_fill(ws, row, col=14, background_color=rebeccaPurpleFill)

        ws.cell(row=row, column=15).value = metric.slas_met
        ws.cell(row=row, column=16).value = metric.delays_introduced_time
        ws.cell(row=row, column=17).value = metric.uat_defects_not_prevented
        ws.cell(row=row, column=18).value = metric.sdis_not_prevented
        ws.cell(row=row, column=19).value = metric.resource_swap
        ws.cell(row=row, column=20).value = metric.escalations
        ws.cell(row=row, column=21).value = metric.rework_introduced_time

        # column 22
        background_color_fill(ws, row, col=22, background_color=saddleBrownFill)

        ws.cell(row=row, column=23).value = metric.avg_team_size
        ws.cell(row=row, column=24).value = metric.overtime_weekday
        ws.cell(row=row, column=25).value = metric.overtime_weekend
        ws.cell(row=row, column=26).value = metric.avg_throughput
        ws.cell(row=row, column=27).value = metric.rework_time
        ws.cell(row=row, column=28).value = metric.resource_swap_time

        # column 29
        background_color_fill(ws, row, col=29, background_color=oliveDrabFill)

        # add $ symbol
        add_dollar_symbol(ws, row, col_start=30, col_end=32)

        ws.cell(row=row, column=30).value = metric.operational_cost
        ws.cell(row=row, column=31).value = metric.license_cost
        ws.cell(row=row, column=32).value = metric.total_operational_cost

        # column 33
        background_color_fill(ws, row, col=33, background_color=darkGreenFill)

        ws.cell(row=row, column=34).value = metric.pheme_manual_tests
        ws.cell(row=row, column=35).value = metric.pheme_auto_tests
        ws.cell(row=row, column=36).value = metric.visilog_txl_parsed
        ws.cell(row=row, column=37).value = metric.visilog_txl_schema_violation

        add_dollar_symbol(ws, row, col_start=38, col_end=38)
        ws.cell(row=row, column=38).value = metric.other_savings

    elif metric.functional_group.key == 'TL':
        # write head title
        write_head_title(ws, HEAD_TL)

        row = 4

        # apply border style
        apply_border_style(ws, row + 1, len(HEAD_TL) + 1)

        # column 1-6
        write_human_resource(ws, row, metric)

        ws.cell(row=row, column=7).value = metric.tickets_received
        ws.cell(row=row, column=8).value = metric.tickets_closed
        ws.cell(row=row, column=9).value = metric.virtual_machines
        ws.cell(row=row, column=10).value = metric.physical_machines

        # column 11
        background_color_fill(ws, row, col=11, background_color=rebeccaPurpleFill)

        ws.cell(row=row, column=12).value = metric.power_consumption_ups_a
        ws.cell(row=row, column=13).value = metric.power_consumption_ups_b
        ws.cell(row=row, column=14).value = metric.license_cost

        add_dollar_symbol(ws, row, col_start=14, col_end=14)

    elif metric.functional_group.key == 'RE':
        write_head_title(ws, HEAD_RE)

        row = 4
        apply_border_style(ws, row + 1, len(HEAD_RE) + 1)

        # column 1-6
        write_human_resource(ws, row, metric)

        # column 7
        ws.cell(row=row, column=7).value = metric.backlog
        ws.cell(row=row, column=8).value = metric.active_projects
        ws.cell(row=row, column=9).value = metric.team_initiative
        ws.cell(row=row, column=10).value = metric.elicitation_analysis_time

        # column 11
        background_color_fill(ws, row, col=11, background_color=rebeccaPurpleFill)

        # column 12
        ws.cell(row=row, column=12).value = metric.revisions
        ws.cell(row=row, column=13).value = metric.rework_introduced_time
        ws.cell(row=row, column=14).value = metric.avg_throughput
        ws.cell(row=row, column=15).value = metric.slas_met
        ws.cell(row=row, column=16).value = metric.slas_missed
        ws.cell(row=row, column=17).value = metric.delays_introduced_time
        ws.cell(row=row, column=18).value = metric.escalations

        # column 19
        background_color_fill(ws, row, col=19, background_color=saddleBrownFill)

        # column 20
        ws.cell(row=row, column=20).value = metric.gross_available_time
        ws.cell(row=row, column=21).value = metric.efficiency
        ws.cell(row=row, column=21).number_format = '0.00%'
        ws.cell(row=row, column=22).value = metric.overtime_weekend
        ws.cell(row=row, column=23).value = metric.overtime_weekday
        ws.cell(row=row, column=24).value = metric.rework_external_time

        # column 25
        background_color_fill(ws, row, col=25, background_color=oliveDrabFill)

        # column 26
        add_dollar_symbol(ws, row, col_start=26, col_end=26)
        add_dollar_symbol(ws, row, col_start=28, col_end=30)

        ws.cell(row=row, column=26).value = metric.rework_external_cost
        ws.cell(row=row, column=27).value = metric.resource_swap_time
        ws.cell(row=row, column=28).value = metric.operational_cost
        ws.cell(row=row, column=29).value = metric.travel_cost
        ws.cell(row=row, column=30).value = metric.overall_cost

    else:
        head = []
        if metric.functional_group.key == 'PQ':
            head = HEAD_PQ + HEAD_TESTING
        elif metric.functional_group.key == 'QA':
            head = HEAD_QA + HEAD_TESTING
        elif metric.functional_group.key == 'TE':
            head = HEAD_TE + HEAD_TESTING

        write_head_title(ws, head)
        row = 4
        apply_border_style(ws, row + 1, len(head) + 1)

        # column 1-6
        write_human_resource(ws, row, metric)

        # column 7
        ws.cell(row=row, column=7).value = metric.team_initiative
        ws.cell(row=row, column=8).value = metric.ticket_backlog
        ws.cell(row=row, column=9).value = metric.ticket_prep
        ws.cell(row=row, column=10).value = metric.ticket_execution
        ws.cell(row=row, column=11).value = metric.ticket_closed
        ws.cell(row=row, column=12).value = metric.project_backlog
        ws.cell(row=row, column=13).value = metric.project_prep
        ws.cell(row=row, column=14).value = metric.project_execution
        ws.cell(row=row, column=15).value = metric.project_closed
        ws.cell(row=row, column=16).value = metric.tc_manual_dev
        ws.cell(row=row, column=17).value = metric.tc_manual_dev_time
        ws.cell(row=row, column=18).value = metric.tc_auto_dev
        ws.cell(row=row, column=19).value = metric.tc_auto_dev_time

        ws.cell(row=row, column=20).value = metric.auto_footprint_dev_age
        ws.cell(row=row, column=20).number_format = '0.00%'

        ws.cell(row=row, column=21).value = metric.tc_manual_execution
        ws.cell(row=row, column=22).value = metric.tc_manual_execution_time
        ws.cell(row=row, column=23).value = metric.tc_auto_execution
        ws.cell(row=row, column=24).value = metric.tc_auto_execution_time

        ws.cell(row=row, column=25).value = metric.auto_footprint_execution_age
        ws.cell(row=row, column=25).number_format = '0.00%'

        ws.cell(row=row, column=26).value = metric.avg_throughput

        # column 27
        background_color_fill(ws, row, col=27, background_color=rebeccaPurpleFill)

        # column 28
        ws.cell(row=row, column=28).value = metric.slas_met
        ws.cell(row=row, column=29).value = metric.delays_introduced_time
        ws.cell(row=row, column=30).value = metric.defect_caught
        ws.cell(row=row, column=31).value = metric.uat_defects_not_prevented
        ws.cell(row=row, column=32).value = metric.sdis_not_prevented
        ws.cell(row=row, column=33).value = metric.resource_swap
        ws.cell(row=row, column=34).value = metric.escalations
        ws.cell(row=row, column=35).value = metric.standards_violated
        ws.cell(row=row, column=36).value = metric.rework_introduced_time

        # column 37
        background_color_fill(ws, row, col=37, background_color=saddleBrownFill)

        # column 38
        ws.cell(row=row, column=38).value = metric.avg_team_size
        ws.cell(row=row, column=39).value = metric.avg_time_frame
        ws.cell(row=row, column=40).value = metric.active_tickets
        ws.cell(row=row, column=41).value = metric.active_projects
        ws.cell(row=row, column=42).value = metric.auto_and_execution_time
        ws.cell(row=row, column=43).value = metric.gross_available_time

        ws.cell(row=row, column=44).value = metric.efficiency
        ws.cell(row=row, column=44).number_format = '0.00%'

        ws.cell(row=row, column=45).value = metric.overtime_weekend
        ws.cell(row=row, column=46).value = metric.overtime_weekday
        ws.cell(row=row, column=47).value = metric.rework_time
        ws.cell(row=row, column=48).value = metric.resource_swap_time

        # column 49
        background_color_fill(ws, row, col=49, background_color=oliveDrabFill)

        # column 50
        add_dollar_symbol(ws, row, col_start=50, col_end=54)
        ws.cell(row=row, column=50).value = metric.operational_cost
        ws.cell(row=row, column=51).value = metric.license_cost
        ws.cell(row=row, column=52).value = metric.total_operational_cost
        ws.cell(row=row, column=53).value = metric.auto_savings
        ws.cell(row=row, column=54).value = metric.other_savings


def write_head_title(ws, title):
    for col in range(1, len(title) + 1):
        cell = ws.cell(row=1, column=col)
        cell.value = title[col - 1]
        cell.alignment = headAlignment
        cell.font = Font(bold=True)


def apply_border_style(ws, rows, columns):
    for r_index in range(1, rows):
        for c_index in range(1, columns):
            ws.cell(row=r_index, column=c_index).border = thinBorder
            ws.cell(row=r_index, column=c_index).number_format = '0.00'


def write_human_resource(ws, row, metric):
    # column 1 Weed Ending
    ws.cell(row=row, column=1).value = get_week_ending_date(metric)

    # column 2
    background_color_fill(ws, row, col=2, background_color=blackFill)

    ws.cell(row=row, column=3).value = metric.staffs
    ws.cell(row=row, column=4).value = metric.openings
    ws.cell(row=row, column=5).value = metric.contractors

    # column 6
    background_color_fill(ws, row, col=6, background_color=cornFlowerBlueFill)


def get_week_ending_date(metric):
    return str(metric.created.year) + '-' \
               + str(metric.created.month) + '-' \
               + str(metric.created.day)


def background_color_fill(ws, row, col, background_color):
    ws.cell(row=1, column=col).font = Font(color=colors.WHITE)
    for r_index in range(1, row + 1):
        ws.cell(row=r_index, column=col).fill = background_color


def add_dollar_symbol(ws, row, col_start, col_end):
    for r_index in range(2, row + 1):
        for c_index in range(col_start, col_end + 1):
            ws.cell(row=r_index, column=c_index).number_format = '$0.00'


