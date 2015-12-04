from datetime import datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.template import RequestContext
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from scorecard.apps.exportations.utils import write_to_excel, write_to_excel_all

from scorecard.apps.users.models import FunctionalGroup
from scorecard.apps.teams.models import InnovationMetrics, LabMetrics, RequirementMetrics, TestMetrics


@login_required
def exportations(request):
    functional_groups = FunctionalGroup.objects.all().order_by('name')
    data = {}
    data_name = []
    data_key = []
    for functional_group in functional_groups:
        data_name.append(functional_group.name)
        data_key.append(functional_group.key)

    dates = InnovationMetrics.objects.values_list('created', flat=True)
    data['name'] = data_name
    data['key'] = data_key
    data['dates'] = dates

    context = RequestContext(request, {
        'data': data
    })

    return render(request, 'exportations/exportations.html', context)


@login_required
def export_excel(request):
    wb = Workbook()
    wb.active.title = 'Testing Summary'
    functional_groups = FunctionalGroup.objects.all().order_by('name')
    key = request.GET.get('key', '')
    date = request.GET.get('date', '')

    update_error = False
    update_error_list = []

    if date:
        date = datetime.strptime(date, '%b. %d, %Y')
        # print date

        if key == 'SU':
            for functional_group in functional_groups:
                ws = wb.create_sheet(functional_group.name)
                if functional_group.key == 'RE':
                    metric = functional_group.requirementmetrics_set.get(created__year=date.year,
                                                                         created__month=date.month,
                                                                         created__day=date.day)
                    if not metric.updated:
                        update_error = True
                        update_error_list.append(str(functional_group.key))

                elif functional_group.key == 'TL':
                    metric = functional_group.labmetrics_set.get(created__year=date.year,
                                                                 created__month=date.month,
                                                                 created__day=date.day)
                    if not metric.updated:
                        update_error = True
                        update_error_list.append(str(functional_group.key))

                elif functional_group.key == 'QI':
                    metric = functional_group.innovationmetrics_set.get(created__year=date.year,
                                                                        created__month=date.month,
                                                                        created__day=date.day)
                    if not metric.updated:
                        update_error = True
                        update_error_list.append(str(functional_group.key))

                else:
                    metric = functional_group.testmetrics_set.get(created__year=date.year,
                                                                  created__month=date.month,
                                                                  created__day=date.day)
                    if not metric.updated:
                        update_error = True
                        update_error_list.append(str(functional_group.key))

                write_to_excel(metric, ws)

            if update_error:
                messages.error(request, 'Team {0} not updated'.format(update_error_list))
                return redirect('exportations:exportations')

        else:
            ws = wb.create_sheet(key)
            # ws.sheet_properties.tabColor = '1072BA'
            if key == 'RE':
                metric = RequirementMetrics.objects.get(functional_group__key=key,
                                                        created__year=date.year,
                                                        created__month=date.month,
                                                        created__day=date.day)

            elif key == 'TL':
                metric = LabMetrics.objects.get(functional_group__key=key,
                                                created__year=date.year,
                                                created__month=date.month,
                                                created__day=date.day)

            elif key == 'QI':
                metric = InnovationMetrics.objects.get(functional_group__key=key,
                                                       created__year=date.year,
                                                       created__month=date.month,
                                                       created__day=date.day)

            else:
                metric = TestMetrics.objects.get(functional_group__key=key,
                                                 created__year=date.year,
                                                 created__month=date.month,
                                                 created__day=date.day)

            if not metric.updated:
                messages.error(request, 'Team \'{0}\' not updated'.format(metric.functional_group.name))
                return redirect('exportations:exportations')

            write_to_excel(metric, ws)

    else:
        # dates = InnovationMetrics.objects.values_list('created', flat=True)
        for functional_group in functional_groups:
            ws = wb.create_sheet(functional_group.name)
            if functional_group.key == 'RE':
                metrics = RequirementMetrics.objects.all()
            elif functional_group.key == 'TL':
                metrics = LabMetrics.objects.all()
            elif functional_group.key == 'QI':
                metrics = InnovationMetrics.objects.all()
            else:
                metrics = TestMetrics.objects.filter(functional_group=functional_group)

            write_to_excel_all(metrics, ws, functional_group.key)

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-disposition'] = 'attachment; filename="foo.xlsx"'
    return response
